import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

P = r"C:\Users\pc\Desktop\Emploi\Offres_Emploi_Genie_Industriel_Lean_2026.xlsx"
wb = load_workbook(P)
print("SHEETS:", wb.sheetnames)
print()
for n in wb.sheetnames:
    ws = wb[n]
    print(f"{n:26s} {ws.dimensions:12s} freeze={str(ws.freeze_panes):6s} filter={ws.auto_filter.ref}")

ws = wb["MAROC - JUNIOR"]
print("\ncolonnes:", [c.value for c in ws[2]])
print("\nligne 3 :", [str(ws.cell(row=3, column=c).value or "")[:26] for c in range(1, 9)])
links = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=13).hyperlink)
print(f"\nliens cliquables MAROC-JUNIOR : {links}/{ws.max_row - 2}")
bad = sum(1 for n in wb.sheetnames for r in range(3, wb[n].max_row + 1)
          if "technicien" in str(wb[n].cell(row=r, column=2).value or "").lower()
          or "technician" in str(wb[n].cell(row=r, column=2).value or "").lower())
print(f"titres 'technicien' restants : {bad}")
