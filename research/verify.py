"""Verification du classeur livre - avant / apres.

    python research/verify.py [<classeur.xlsx>] [<classeur_avant.xlsx>]

Controle ce qui a ete demande, et rien d autre: repartition des feuilles, aucune
offre fermee, aucune ligne junior dont le plancher d experience est positif, liens
cliquables et distincts, cellules vides par colonne avant/apres."""
import sys, io, os, re, glob, hashlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "skill", "findmyprincessajob", "pipeline"))
from openpyxl import load_workbook
from build import exp_floor, CLOSED_RE, TICK
from core import strip_accents

XL = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    ROOT, "Offres_Emploi_Genie_Industriel_Lean_2026.xlsx")
if len(sys.argv) > 2:
    OLD = sys.argv[2]
else:
    cands = sorted(glob.glob(os.path.join(ROOT, "backup", "*avant_round4*.xlsx")))
    OLD = cands[-1] if cands else None

LINK_COL = "LIEN DE L'OFFRE"


def read(path):
    wb = load_workbook(path)
    out, per = [], {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr = [c.value for c in ws[2]]
        li = hdr.index(LINK_COL) + 1 if LINK_COL in hdr else None
        n = 0
        for r in range(3, ws.max_row + 1):
            vals = [c.value for c in ws[r]]
            if not any(vals):
                continue
            d = {h: (str(v).strip() if v is not None else "") for h, v in zip(hdr, vals)}
            d["_sheet"], d["_row"] = sn, r
            d["_link"] = bool(li and ws.cell(row=r, column=li).hyperlink)
            out.append(d)
            n += 1
        per[sn] = n
    return out, per, hdr


new, per, hdr = read(XL)
old, perold, hdrold = read(OLD) if OLD and os.path.exists(OLD) else ([], {}, [])

print("=" * 74)
print(f"CLASSEUR : {os.path.basename(XL)}")
if old:
    print(f"REFERENCE: {os.path.basename(OLD)}")
print("=" * 74)

print("\n1. LIGNES PAR FEUILLE")
names = list(dict.fromkeys(list(perold) + list(per)))
for sn in names:
    a, b = perold.get(sn, 0), per.get(sn, 0)
    print(f"   {sn:30s} {a:5d} -> {b:5d}   ({b - a:+d})")
print(f"   {'TOTAL':30s} {sum(perold.values()):5d} -> {sum(per.values()):5d}"
      f"   ({sum(per.values()) - sum(perold.values()):+d})")

print("\n2. OFFRES FERMEES")
closed = [r for r in new if CLOSED_RE.search(strip_accents(
    r.get("Date limite candidature", "") + " " + r.get("Date de publication", "")))]
print(f"   attendu 0 -> trouve {len(closed)}   {'OK' if not closed else 'ECHEC'}")
for r in closed[:10]:
    print(f"      [{r['_sheet'][:6]} L{r['_row']}] {r.get('Titre du poste', '')[:60]}")

print("\n3. FEUILLE JUNIOR : PLANCHER D EXPERIENCE")
jun = [r for r in new if "JUNIOR" in r["_sheet"]]
bad = [(r, exp_floor(r.get("Expérience requise"))) for r in jun]
bad = [(r, f) for r, f in bad if f is not None and f > 0]
print(f"   lignes junior: {len(jun)}   plancher > 0: {len(bad)}   "
      f"{'OK' if not bad else 'ECHEC'}")
for r, f in bad[:15]:
    print(f"      [L{r['_row']}] plancher={f} [{r.get('Expérience requise')}] "
          f"{r.get('Titre du poste', '')[:45]}")
zero = sum(1 for r in jun if exp_floor(r.get("Expérience requise")) == 0)
sil = sum(1 for r in jun if exp_floor(r.get("Expérience requise")) is None)
print(f"   dont plancher explicitement zero: {zero} | annonce muette (plancher "
      f"inconnu, non nul prouve): {sil}")

print("\n4. LIENS")
urls = [r.get(LINK_COL, "") for r in new]
clic = sum(1 for r in new if r["_link"])
print(f"   cliquables: {clic}/{len(new)}   {'OK' if clic == len(new) else 'ECHEC'}")
print(f"   distincts : {len(set(urls))}/{len(new)}   "
      f"{'OK' if len(set(urls)) == len(new) else 'ECHEC'}")
dups = {u for u in urls if urls.count(u) > 1}
for u in list(dups)[:5]:
    print("      doublon:", u)

print("\n5. CELLULES VIDES PAR COLONNE")
cols = [h for h in hdr if h]
w = max(len(c) for c in cols)
print(f"   {'colonne':{w}s}  avant  apres   delta   remplissage")
for c in cols:
    # the evidence column only exists on the REMOTE sheets; counting it against
    # every Moroccan row would report a 99 % hole that is not one
    scope = [r for r in new if r["_sheet"].startswith("REMOTE")] if "Preuve" in c else new
    scopeo = [r for r in old if r["_sheet"].startswith("REMOTE")] if "Preuve" in c else old
    b = sum(1 for r in scope if not r.get(c))
    a = sum(1 for r in scopeo if not r.get(c)) if old and c in hdrold else None
    pct = 100 * (len(scope) - b) / max(1, len(scope))
    astr = f"{a:5d}" if a is not None else "    -"
    dstr = f"{b - a:+6d}" if a is not None else "     -"
    note = f"   (sur {len(scope)} lignes REMOTE)" if "Preuve" in c else ""
    print(f"   {c:{w}s}  {astr}  {b:5d}  {dstr}   {pct:5.1f}%{note}")

print("\n6. COLONNE 'Postule'")
tick = sum(1 for r in new if TICK in (r.get("Postulé") or ""))
print(f"   presente: {'Postulé' in cols}   1re colonne: {cols[0] == 'Postulé'}   "
      f"deja cochees: {tick}")

print("\n7. COPIE DRIVE")
import json
spec = json.load(open(os.path.join(ROOT, "spec.json"), encoding="utf-8"))
pub = spec.get("publish_dir")
tgt = os.path.join(pub, os.path.basename(XL)) if pub else None


def sha(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 16), b""):
            h.update(b)
    return h.hexdigest()


if tgt and os.path.exists(tgt):
    a, b = sha(XL), sha(tgt)
    print(f"   bureau {a[:16]}")
    print(f"   drive  {b[:16]}   {'IDENTIQUE' if a == b else 'DIFFERENT'}")
    print(f"   -> {tgt}")
else:
    print(f"   copie Drive absente ({tgt})")
