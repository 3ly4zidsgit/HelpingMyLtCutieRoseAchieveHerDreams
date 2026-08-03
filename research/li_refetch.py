"""Re-fetch LinkedIn ad bodies through the guest API.

The public /jobs/view/ page often serves only the login shell: title, company,
"Decouvrez qui X a recrute pour ce poste", then the sign-in form. Every column
that lives in the description then comes out empty for no reason other than which
variant of the page answered. /jobs-guest/jobs/api/jobPosting/<id> returns the
description itself and needs no auth (see references/SOURCES.md).

    python research/li_refetch.py            # every LinkedIn row of the workbook
"""
import sys, io, os, re, json, time, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "skill", "findmyprincessajob", "pipeline"))
import requests
from core import HDRS, jid
from openpyxl import load_workbook

FT = os.path.join(ROOT, "data", "fulltext.json")
full = json.load(open(FT, encoding="utf-8")) if os.path.exists(FT) else {}

# no account, no cookie: the guest endpoint is public. Nothing here can get the
# user's LinkedIn banned, which is why it is the only route this pipeline uses.
WORKBOOK_ONLY = "--all" not in sys.argv
XL = os.path.join(ROOT, "Offres_Emploi_Genie_Industriel_Lean_2026.xlsx")
urls = []
wb = load_workbook(XL)
for sn in wb.sheetnames:
    ws = wb[sn]
    hdr = [c.value for c in ws[2]]
    if "LIEN DE L'OFFRE" not in hdr:
        continue
    i = hdr.index("LIEN DE L'OFFRE")
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=i + 1).value
        if v and "linkedin.com/jobs/view" in str(v):
            urls.append(str(v).strip())
if not WORKBOOK_ONLY:
    for p in ("raw_http", "raw_uc", "raw_extra"):
        pp = os.path.join(ROOT, "data", p + ".json")
        if os.path.exists(pp):
            for row in json.load(open(pp, encoding="utf-8")):
                u = (row.get("url") or "")
                if "linkedin.com/jobs/view" in u:
                    urls.append(u)
urls = list(dict.fromkeys(urls))

TAG = re.compile(r"<[^>]+>")
WALL = re.compile(r"Inscrivez-vous pour postuler|Mot de passe oubli|S.identifier avec")


def clean(html):
    t = re.sub(r"(?is)<(script|style).*?</\1>", " ", html)
    t = re.sub(r"(?i)<br\s*/?>|</p>|</li>|</div>", "\n", t)
    t = TAG.sub(" ", t)
    for a, b in (("&nbsp;", " "), ("&amp;", "&"), ("&#39;", "'"), ("&quot;", '"'),
                 ("&lt;", "<"), ("&gt;", ">"), ("&eacute;", "é"), ("&egrave;", "è")):
        t = t.replace(a, b)
    return re.sub(r"\n{2,}", "\n", re.sub(r"[ \t]+", " ", t)).strip()


def usable(u):
    """Does the stored body actually contain the ad, or only the login wall?

    Length is no test: the wall alone runs past 2 000 characters of cookie policy
    and "Identifiez-vous pour acceder a des conseils generes par l IA". What marks
    a real description is the marker LinkedIn puts around it."""
    d = full.get(u) or {}
    if d.get("src") == "linkedin-guest-api":
        return True
    t = d.get("text", "")
    return "Show more" in t and len(WALL.sub("", t)) > 2500


todo = [u for u in urls if not usable(u)]
print(f"{len(urls)} annonces LinkedIn, {len(todo)} sans corps exploitable")
s = requests.Session()
s.headers.update(HDRS)
ok = fail = 0
for i, u in enumerate(todo, 1):
    m = re.search(r"(\d{6,})(?:\?|$|/)", u)
    if not m:
        fail += 1
        continue
    api = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{m.group(1)}"
    try:
        r = s.get(api, timeout=25)
        if r.status_code != 200 or len(r.text) < 500:
            fail += 1
        else:
            txt = clean(r.text)
            prev = (full.get(u) or {}).get("text", "")
            if len(txt) > 400:
                full[u] = {"text": (prev + "\n" + txt) if prev else txt,
                           "title": "", "ok": "True", "src": "linkedin-guest-api"}
                ok += 1
            else:
                fail += 1
    except Exception:
        fail += 1
    if i % 20 == 0 or i == len(todo):
        print(f"  {i}/{len(todo)}  ok={ok} echec={fail}", flush=True)
        json.dump(full, open(FT, "w", encoding="utf-8"), ensure_ascii=False)
    time.sleep(random.uniform(1.1, 2.4))
json.dump(full, open(FT, "w", encoding="utf-8"), ensure_ascii=False)
print(f"termine: {ok} corps recuperes, {fail} echecs -> data/fulltext.json ({len(full)} entrees)")
