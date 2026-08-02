import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

P = r"C:\Users\pc\Desktop\Emploi\Offres_Emploi_Genie_Industriel_Lean_2026_NOUVEAU.xlsx"
wb = load_workbook(P)
print("SHEETS:", wb.sheetnames, "\n")
for n in wb.sheetnames:
    ws = wb[n]
    print(f"{n:26s} {ws.dimensions:12s} filter={ws.auto_filter.ref}")

ws = wb["MAROC - JUNIOR"]
hdr = [c.value for c in ws[2]]
print(f"\n{len(hdr)} COLONNES : {hdr}")

bad = 0
for n in wb.sheetnames[1:]:
    w = wb[n]
    for r in range(3, w.max_row + 1):
        t = str(w.cell(row=r, column=2).value or "").lower()
        if "technicien" in t or "technician" in t:
            bad += 1
print(f"\ntitres contenant 'technicien' : {bad}")

links = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=13).hyperlink)
print(f"liens cliquables (MAROC-JUNIOR) : {links}/{ws.max_row - 2}")

print("\n--- MAROC - JUNIOR (8 premiers) ---")
for r in range(3, 11):
    print("  " + " | ".join(str(ws.cell(row=r, column=c).value or "")[:24] for c in (2, 3, 4, 6, 7, 11)))

we = wb["MAROC - AVEC EXPERIENCE"]
print("\n--- MAROC - AVEC EXPERIENCE (6 premiers) ---")
for r in range(3, 9):
    print("  " + " | ".join(str(we.cell(row=r, column=c).value or "")[:24] for c in (2, 3, 4, 6, 7, 11)))

for sh in ("REMOTE - JUNIOR", "REMOTE - AVEC EXPERIENCE"):
    w = wb[sh]
    print(f"\n--- {sh} ({w.max_row - 2}) ---")
    print("   colonnes:", len([c.value for c in w[2]]))
    for r in range(3, w.max_row + 1):
        print(f"  {str(w.cell(row=r, column=2).value)[:44]:44s} | {str(w.cell(row=r, column=3).value)[:18]:18s} "
              f"| {str(w.cell(row=r, column=4).value)[:12]:12s} | {str(w.cell(row=r, column=15).value)[:70]}")
