"""Audit du classeur: lignes par feuille, cellules vides par colonne, defauts."""
import sys, io, os, re, json, hashlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

XL = sys.argv[1] if len(sys.argv) > 1 else \
    r"c:\Users\El Yazid\Desktop\HelpingMyLtCutieRoseAchieveHerDreams\Offres_Emploi_Genie_Industriel_Lean_2026.xlsx"

wb = load_workbook(XL)
total = 0
allrows = []
for sn in wb.sheetnames:
    ws = wb[sn]
    hdr = [c.value for c in ws[2]]
    n = 0
    for r in range(3, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        if not any(vals):
            continue
        d = {h: (str(v).strip() if v is not None else "") for h, v in zip(hdr, vals)}
        d["_sheet"] = sn
        d["_row"] = r
        d["_link"] = ws.cell(row=r, column=hdr.index("LIEN DE L'OFFRE") + 1).hyperlink is not None
        allrows.append(d)
        n += 1
    print(f"{sn:32s} {n:5d}")
    total += n
print(f"{'TOTAL':32s} {total:5d}\n")

cols = [c.value for c in wb[wb.sheetnames[0]][2]]
print("cellules vides par colonne (sur %d):" % total)
for c in cols:
    e = sum(1 for r in allrows if not r.get(c))
    if e:
        print(f"  {e:5d}  {c}")

print("\n--- defauts ---")
exp = [r for r in allrows if "EXPIREE" in (r.get("Date limite candidature") or "").upper()]
print(f"OFFRE EXPIREE               : {len(exp)}")

FLOOR = re.compile(r"\b(\d+(?:[.,]\d+)?)\s*(ans?|an|mois|années?|years?|month)", re.I)
def floor_pos(s):
    """plancher > 0 ?"""
    t = (s or "").lower()
    if not t:
        return None
    if re.search(r"d[ée]butant|jeune diplom|sortie d.?[ée]tudes|sans exp[ée]rience|"
                 r"fra[iî]chement diplom|entry.level|premier emploi|\bpfe\b|^0$", t):
        return False
    m = re.search(r"\b0\s*[-aà]\s*\d", t)
    if m:
        return False
    m = FLOOR.search(t)
    if m:
        try:
            return float(m.group(1).replace(",", ".")) > 0
        except ValueError:
            return None
    return None

jun = [r for r in allrows if "JUNIOR" in r["_sheet"]]
bad = [r for r in jun if floor_pos(r.get("Expérience requise")) is True]
print(f"JUNIOR avec plancher > 0    : {len(bad)}")
for r in bad:
    print(f"    [{r['_sheet'][:6]} L{r['_row']}] {r.get('Expérience requise'):28s} | "
          f"{(r.get('Titre du poste') or '')[:52]}")

urls = [r.get("LIEN DE L'OFFRE") for r in allrows]
print(f"\nliens cliquables            : {sum(1 for r in allrows if r['_link'])}/{total}")
print(f"liens distincts             : {len(set(urls))}/{total}")
dups = {u for u in urls if urls.count(u) > 1}
if dups:
    print(f"  doublons: {len(dups)}")
    for u in list(dups)[:10]:
        print("   ", u)

json.dump(allrows, open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
         "..", "data", "audit_rows.json"), "w", encoding="utf-8"),
         ensure_ascii=False, indent=1)
