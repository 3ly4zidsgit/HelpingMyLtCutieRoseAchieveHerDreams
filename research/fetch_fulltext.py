"""Re-fetch the FULL text of every offer in the workbook.

The round-1/2 scrapers kept only what the list page showed (~700 chars), which is
why 58 % of 'Experience requise' and 68 % of 'Type de contrat' are empty. The
requirement is almost always in the body of the ad, not the listing card.

Plain HTTP works for LinkedIn (jobs-guest needs no auth), Rekrute and
SmartRecruiters. Everything else answers 403 and needs UC mode - see
fetch_fulltext_uc.py.

Writes data/fulltext.json as {url: {"text": ..., "title": ..., "ok": bool}}.
Resumable: an existing entry is never refetched.
"""
import sys, io, os, re, json, time, html, random
import urllib.request, urllib.error

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "fulltext.json")

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36")
HDRS = {"User-Agent": UA,
        "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "identity"}

PLAIN = re.compile(r"(linkedin\.com|rekrute\.com|smartrecruiters\.com)", re.I)


def clean(h):
    h = re.sub(r"<(script|style|noscript)[^>]*>.*?</\1>", " ", h, flags=re.S | re.I)
    h = re.sub(r"<br\s*/?>|</p>|</li>|</div>|</tr>|</h\d>", "\n", h, flags=re.I)
    h = re.sub(r"<li[^>]*>", " - ", h, flags=re.I)
    h = re.sub(r"<[^>]+>", " ", h)
    h = html.unescape(h)
    h = re.sub(r"[ \t\xa0]+", " ", h)
    return re.sub(r"\n\s*\n+", "\n", h).strip()


def get(url, timeout=35):
    req = urllib.request.Request(url, headers=HDRS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace")


def fetch_one(url):
    """Return (text, title). LinkedIn goes through the guest posting endpoint,
    which returns the full description without auth."""
    m = re.search(r"linkedin\.com/jobs/view/(?:.*-)?(\d{6,})", url)
    if m:
        api = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{m.group(1)}"
        h = get(api)
        t = clean(h)
        return t, ""
    h = get(url)
    ttl = re.search(r"<title[^>]*>(.*?)</title>", h, re.S | re.I)
    return clean(h), html.unescape(ttl.group(1)).strip() if ttl else ""


def workbook_urls():
    """The workbook is the source of truth: it also holds rows recovered from an
    earlier run that never went through merged.json."""
    from openpyxl import load_workbook
    path = os.path.join(ROOT, "Offres_Emploi_Genie_Industriel_Lean_2026.xlsx")
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    out = []
    for s in wb.sheetnames:
        ws = wb[s]
        hdr = [c.value for c in ws[2]]
        if "LIEN DE L'OFFRE" not in hdr:
            continue
        iu = hdr.index("LIEN DE L'OFFRE")
        for r in range(3, ws.max_row + 1):
            c = [x.value for x in ws[r]]
            if any(c):
                out.append(str(c[iu] or "").strip())
    return out


def main():
    rows = json.load(open(os.path.join(ROOT, "data", "merged.json"), encoding="utf-8"))
    urls = []
    seen = set()
    # workbook first - those are the rows that actually ship
    for u in workbook_urls() + [(r.get("url") or "").strip() for r in rows]:
        if u and u not in seen and PLAIN.search(u):
            seen.add(u)
            urls.append(u)

    store = {}
    if os.path.exists(OUT):
        store = json.load(open(OUT, encoding="utf-8"))
    todo = [u for u in urls if u not in store]
    print(f"{len(urls)} URL accessibles en HTTP simple, {len(todo)} a recuperer", flush=True)

    ok = fail = 0
    for i, u in enumerate(todo, 1):
        try:
            t, ttl = fetch_one(u)
            store[u] = {"text": t, "title": ttl, "ok": len(t) > 400}
            ok += 1
        except urllib.error.HTTPError as e:
            store[u] = {"text": "", "title": "", "ok": False, "err": f"HTTP {e.code}"}
            fail += 1
        except Exception as e:
            store[u] = {"text": "", "title": "", "ok": False, "err": type(e).__name__}
            fail += 1
        if i % 10 == 0 or i == len(todo):
            json.dump(store, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
            print(f"  {i}/{len(todo)}  ok={ok} echec={fail}", flush=True)
        time.sleep(random.uniform(1.1, 2.3))

    json.dump(store, open(OUT, "w", encoding="utf-8"), ensure_ascii=False)
    good = sum(1 for v in store.values() if v.get("ok"))
    print(f"\n{good}/{len(store)} textes complets -> data/fulltext.json")


if __name__ == "__main__":
    main()
