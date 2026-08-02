import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook

P = r"C:\Users\pc\Desktop\Emploi\Offres_Emploi_Genie_Industriel_Lean_2026.xlsx"
wb = load_workbook(P)
print("SHEETS:", wb.sheetnames, "\n")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"{name:26s} dims={ws.dimensions:12s} freeze={str(ws.freeze_panes):6s} "
          f"filter={ws.auto_filter.ref}")
print()

ws = wb["TOUTES LES OFFRES"]
print("NOTE:", ws["A1"].value)
hdr = [c.value for c in ws[2]]
print(f"\n{len(hdr)} COLUMNS:")
for i, h in enumerate(hdr, 1):
    print(f"  {i:2d}. {h}")

print("\n--- first 6 data rows (key columns) ---")
idx = {h: i + 1 for i, h in enumerate(hdr)}
keys = ["Titre du poste", "Entreprise", "Type de contrat", "Niveau d'expérience",
        "Ville", "Pays", "Date de publication", "Date limite candidature",
        "Contact RH / Recruteur", "Email de candidature", "Email société (générique)", "Source"]
for r in range(3, 9):
    vals = [str(ws.cell(row=r, column=idx[k]).value or "")[:22] for k in keys]
    print("  " + " | ".join(vals))

link_col = idx["LIEN DE L'OFFRE"]
live = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=link_col).hyperlink)
print(f"\nclickable offer links: {live}/{ws.max_row - 2}")
mail_col = idx["Email société (générique)"]
ml = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=mail_col).hyperlink)
print(f"clickable mailto links: {ml}")

ab = wb["HORS MAROC (REMOTE)"]
print("\n--- HORS MAROC sample ---")
for r in range(3, 11):
    print("  ", str(ab.cell(row=r, column=idx["Titre du poste"]).value)[:44],
          "|", str(ab.cell(row=r, column=idx["Entreprise"]).value)[:22],
          "|", str(ab.cell(row=r, column=idx["Pays"]).value)[:22],
          "|", str(ab.cell(row=r, column=idx["Télétravail / Remote"]).value)[:12])

jr = wb["JUNIOR - SANS EXPERIENCE"]
print("\n--- JUNIOR sample ---")
for r in range(3, 11):
    print("  ", str(jr.cell(row=r, column=idx["Titre du poste"]).value)[:46],
          "|", str(jr.cell(row=r, column=idx["Entreprise"]).value)[:20],
          "|", str(jr.cell(row=r, column=idx["Expérience requise"]).value)[:18],
          "|", str(jr.cell(row=r, column=idx["Ville"]).value)[:14])
