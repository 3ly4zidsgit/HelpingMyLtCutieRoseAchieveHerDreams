"""Final workbook: 4 sheets (Maroc/Junior, Maroc/Experimente, Remote/Junior,
Remote/Experimente) + SYNTHESE. Reduced column set, no 'technicien' roles, and
remote offers restricted to genuinely worldwide / visa-free positions."""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from common import load, norm

DEST = r"C:\Users\pc\Desktop\Emploi\Offres_Emploi_Genie_Industriel_Lean_2026.xlsx"

COLS = [
    ("N°",                      "idx",                  6),
    ("Titre du poste",          "job_title",            48),
    ("Entreprise",              "company",              28),
    ("Type de contrat",         "contract_type",        18),
    ("Niveau d'expérience",     "seniority_bucket",     19),
    ("Expérience requise",      "experience_required",  22),
    ("Ville",                   "location_city",        22),
    ("Pays / Éligibilité",      "country",              24),
    ("Secteur",                 "sector",               26),
    ("Fonction",                "function",             26),
    ("Date de publication",     "date_posted",          18),
    ("Date limite candidature", "deadline",             18),
    ("LIEN DE L'OFFRE",         "url",                  54),
    ("Description (extrait)",   "description_snippet",  90),
]
# extra column shown only on the two remote sheets
REMOTE_COL = ("Pourquoi 100% remote / sans visa", "remote_reason", 60)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="EEF3FA")
LINK_FONT = Font(color="0563C1", underline="single", size=10)
CELL_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
GREY = PatternFill("solid", fgColor="F2F2F2")
BUCKET_FILL = {"Junior / Debutant": GREEN, "Experimente": ORANGE, "Non precise": GREY}

# --------------------------------------------------------------- filters
BAD_TITLE = re.compile(r"\btechnicien(ne)?s?\b|\btechnician[s]?\b")

TXT_GLOBAL = re.compile(r"work from anywhere|anywhere in the world|from any country|"
                        r"any time ?zone|globally distributed|fully distributed|"
                        r"location[- ]independent|work remotely from anywhere")
GEO_GLOBAL = re.compile(r"worldwide|anywhere|global|emea|latam|apac|"
                        r"remote mondial|multi[- ]region")
BLOCKER = re.compile(
    r"must (be (located|based|physically|resident)|reside|live) (in|within)|"
    r"(authoriz|authoris)(ed|ation) to work|right to work in|eligible to work in|"
    r"work permit|visa sponsorship|cannot sponsor|no sponsorship|sponsorship (is )?not|"
    r"security clearance|u\.?s\.? citizen|green card|"
    r"\bhybrid\b|\bon[- ]?site\b|\bonsite\b|\bin[- ]office\b|relocat")

def truly_worldwide(r):
    """fully remote AND open beyond a single country AND no immigration barrier"""
    geo = norm(r.get("location_city", "") + " " + r.get("country", ""))
    txt = norm(r.get("description_snippet", ""))
    if BLOCKER.search(txt):
        return False
    return bool(GEO_GLOBAL.search(geo) or TXT_GLOBAL.search(txt))

def clean(v):
    if v is None: return ""
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", str(v))
    return re.sub(r"\s+", " ", s).strip()

def sort_key(r):
    ts = 0.0
    d = r.get("date_posted_iso") or ""
    if d:
        try: ts = datetime.strptime(d, "%Y-%m-%d").timestamp()
        except ValueError: ts = 0.0
    return (0 if ts else 1, -ts, -r.get("score", 0))

def write_sheet(wb, name, rows, note, cols=None):
    cols = cols or COLS
    ws = wb.create_sheet(name[:31])
    ws.cell(row=1, column=1, value=note).font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.row_dimensions[1].height = 24
    for c, (label, _, w) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 32
    for i, row in enumerate(rows, 1):
        rr = 2 + i
        for c, (label, key, _) in enumerate(cols, 1):
            val = i if key == "idx" else clean(row.get(key, ""))
            cell = ws.cell(row=rr, column=c, value=val)
            cell.font = CELL_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=key in ("job_title", "description_snippet", "sector",
                                                         "function", "remote_reason"),
                                       horizontal="center" if key == "idx" else "left")
            if i % 2 == 0 and key != "seniority_bucket":
                cell.fill = ALT_FILL
            if key == "seniority_bucket":
                cell.fill = BUCKET_FILL.get(row.get("seniority_bucket"), GREY)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if key == "url" and str(val).startswith("http"):
                cell.hyperlink = str(val); cell.font = LINK_FONT
        ws.row_dimensions[rr].height = 30
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(rows)}"
    # freeze the header rows only - no pinned columns
    ws.freeze_panes = ws.cell(row=3, column=1)
    return ws


def main():
    base = load("merged")
    ww = load("worldwide")

    # 1. drop "technicien" everywhere
    before = len(base) + len(ww)
    base = [r for r in base if not BAD_TITLE.search(norm(r.get("job_title", "")))]
    ww = [r for r in ww if not BAD_TITLE.search(norm(r.get("job_title", "")))]
    print(f"'technicien' removed: {before - len(base) - len(ww)}")

    # 2. Morocco
    maroc = [r for r in base if r["country"] == "Maroc"]

    # 3. remote: only offers individually read and confirmed worldwide / visa-free
    from remote_curation import keep_reason
    intl = [r for r in base if r["country"] != "Maroc"]
    remote = {}
    for r in intl + ww:
        reason = keep_reason(r)
        if not reason:
            continue
        k = norm(r.get("job_title", "")) + "|" + norm(r.get("company", ""))
        if k in remote:
            continue
        r["remote_reason"] = reason
        r["country"] = "Mondial - sans visa"
        remote[k] = r
    remote = list(remote.values())
    for r in remote:
        if not r.get("seniority_bucket"):
            r["seniority_bucket"] = "Non precise"
    print(f"international offers read: {len(intl) + len(ww)} -> {len(remote)} confirmed fully remote & visa-free")

    def split(rs):
        jr = [r for r in rs if r["seniority_bucket"] in ("Junior / Debutant", "Non precise")]
        ex = [r for r in rs if r["seniority_bucket"] == "Experimente"]
        return sorted(jr, key=sort_key), sorted(ex, key=sort_key)

    ma_jr, ma_ex = split(maroc)
    re_jr, re_ex = split(remote)

    wb = Workbook(); wb.remove(wb.active)

    write_sheet(wb, "MAROC - JUNIOR", ma_jr,
                f"MAROC — JUNIOR / SANS EXPÉRIENCE ({len(ma_jr)} offres). Expérience non exigée ou non précisée.")
    write_sheet(wb, "MAROC - AVEC EXPERIENCE", ma_ex,
                f"MAROC — AVEC EXPÉRIENCE ({len(ma_ex)} offres). 2 ans et plus / senior / manager / responsable.")
    RCOLS = COLS + [REMOTE_COL]
    write_sheet(wb, "REMOTE - JUNIOR", re_jr,
                f"REMOTE 100% MONDIAL — JUNIOR ({len(re_jr)} offres). Chaque annonce a été lue une par une : "
                "aucune condition de résidence, de visa ni de présence sur site.", RCOLS)
    write_sheet(wb, "REMOTE - AVEC EXPERIENCE", re_ex,
                f"REMOTE 100% MONDIAL — AVEC EXPÉRIENCE ({len(re_ex)} offres). Chaque annonce a été lue une par une : "
                "aucune condition de résidence, de visa ni de présence sur site.", RCOLS)

    dest = DEST
    try:
        wb.save(dest)
    except PermissionError:
        # the workbook is open in Excel; never clobber it, write beside it instead
        dest = DEST.replace(".xlsx", "_NOUVEAU.xlsx")
        wb.save(dest)
        print("\n*** Le fichier d'origine est ouvert dans Excel : ferme-le puis relance,")
        print(f"*** ou utilise directement : {dest}")
    print(f"\nSAVED: {dest}")
    print(f"  MAROC - JUNIOR             {len(ma_jr)}")
    print(f"  MAROC - AVEC EXPERIENCE    {len(ma_ex)}")
    print(f"  REMOTE - JUNIOR            {len(re_jr)}")
    print(f"  REMOTE - AVEC EXPERIENCE   {len(re_ex)}")
    print(f"  TOTAL                      {len(ma_jr)+len(ma_ex)+len(re_jr)+len(re_ex)}")
    print(f"  size: {os.path.getsize(dest)/1024:.1f} KB")

if __name__ == "__main__":
    main()
