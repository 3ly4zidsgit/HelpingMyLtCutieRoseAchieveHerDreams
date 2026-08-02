"""Build the multi-sheet Excel deliverable."""
import os, re, json
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from common import load, norm, OUT

DEST = r"C:\Users\pc\Desktop\Emploi\Offres_Emploi_Genie_Industriel_Lean_2026.xlsx"

COLS = [
    ("N°",                         "idx",                  6),
    ("Titre du poste",             "job_title",            46),
    ("Entreprise",                 "company",              28),
    ("Type de contrat",            "contract_type",        18),
    ("Niveau d'expérience",        "seniority_bucket",     19),
    ("Expérience requise",         "experience_required",  22),
    ("Niveau d'études",            "education_level",      18),
    ("Ville",                      "location_city",        20),
    ("Pays",                       "country",              16),
    ("Télétravail / Remote",       "remote",               18),
    ("Secteur",                    "sector",               24),
    ("Fonction",                   "function",             24),
    ("Nombre de postes",           "positions",            10),
    ("Salaire",                    "salary",               18),
    ("Date de publication",        "date_posted",          18),
    ("Date publication (ISO)",     "date_posted_iso",      15),
    ("Date limite candidature",    "deadline",             16),
    ("Date limite (ISO)",          "deadline_iso",         14),
    ("Contact RH / Recruteur",     "recruiter_or_hr",      26),
    ("Poste du contact RH",        "hr_title",             30),
    ("Profil LinkedIn du contact", "hr_profile",           34),
    ("Email de candidature",       "contact_email",        30),
    ("Email société (générique)",  "company_email",        30),
    ("Site web société",           "company_website",      28),
    ("LIEN DE L'OFFRE",            "url",                  50),
    ("Source",                     "source",               22),
    ("Mots-clés correspondants",   "keywords_matched",     34),
    ("Score pertinence",           "score",                9),
    ("Description (extrait)",      "description_snippet",  80),
]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="EEF3FA")
LINK_FONT = Font(color="0563C1", underline="single", size=10)
CELL_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BUCKET_FILL = {
    "Junior / Debutant": PatternFill("solid", fgColor="C6EFCE"),
    "Experimente":       PatternFill("solid", fgColor="FCE4D6"),
    "Non precise":       PatternFill("solid", fgColor="F2F2F2"),
}

def sort_key(r):
    """Most recently posted first; undated offers fall to the bottom, best score first."""
    ts = 0.0
    d = r.get("date_posted_iso") or ""
    if d:
        try:
            ts = datetime.strptime(d, "%Y-%m-%d").timestamp()
        except ValueError:
            ts = 0.0
    return (0 if ts else 1, -ts, -r.get("score", 0))

def clean(v):
    if v is None: return ""
    s = str(v)
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def write_sheet(wb, name, rows, note=""):
    ws = wb.create_sheet(name[:31])
    r0 = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(bold=True, size=11, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        ws.row_dimensions[1].height = 22
        r0 = 2

    for c, (label, _, w) in enumerate(COLS, 1):
        cell = ws.cell(row=r0, column=c, value=label)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[r0].height = 34

    for i, row in enumerate(rows, 1):
        rr = r0 + i
        for c, (label, key, _) in enumerate(COLS, 1):
            val = i if key == "idx" else row.get(key, "")
            val = val if key == "score" else clean(val)
            cell = ws.cell(row=rr, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=key in ("job_title", "description_snippet",
                                                         "keywords_matched", "hr_title", "sector"),
                                       horizontal="center" if key in ("idx","score","positions") else "left")
            if i % 2 == 0 and key != "seniority_bucket":
                cell.fill = ALT_FILL
            if key == "seniority_bucket":
                cell.fill = BUCKET_FILL.get(row.get("seniority_bucket"), ALT_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if key in ("url", "hr_profile", "company_website") and str(val).startswith("http"):
                cell.hyperlink = str(val); cell.font = LINK_FONT
            if key == "contact_email" and "@" in str(val):
                cell.hyperlink = "mailto:" + str(val).split(",")[0].strip(); cell.font = LINK_FONT
            if key == "company_email" and "@" in str(val):
                cell.hyperlink = "mailto:" + str(val).split(",")[0].strip(); cell.font = LINK_FONT
        ws.row_dimensions[rr].height = 30

    last = r0 + len(rows)
    if rows:
        ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(COLS))}{last}"
    ws.freeze_panes = ws.cell(row=r0 + 1, column=3)
    return ws


def synth_sheet(wb, rows):
    ws = wb.create_sheet("SYNTHESE", 0)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 14

    def block(title, counter, col, top, limit=22):
        c = ws.cell(row=top, column=col, value=title)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HDR_FILL
        c2 = ws.cell(row=top, column=col + 1, value="Offres")
        c2.font = Font(bold=True, color="FFFFFF"); c2.fill = HDR_FILL
        for i, (k, v) in enumerate(counter.most_common(limit), 1):
            ws.cell(row=top + i, column=col, value=str(k) or "(non précisé)").font = CELL_FONT
            ws.cell(row=top + i, column=col + 1, value=v).font = CELL_FONT
        return top + len(counter.most_common(limit)) + 2

    t = ws.cell(row=1, column=1, value="OFFRES D'EMPLOI — GÉNIE INDUSTRIEL / LEAN SIX SIGMA / AMÉLIORATION CONTINUE / EXCELLENCE OPÉRATIONNELLE")
    t.font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value=f"Recherche effectuée le {datetime.now():%d/%m/%Y} — Maroc + postes 100% remote à l'international").font = Font(italic=True, size=11)

    ws.cell(row=4, column=1, value="TOTAL DES OFFRES").font = Font(bold=True, size=12)
    ws.cell(row=4, column=2, value=len(rows)).font = Font(bold=True, size=12, color="C00000")
    ws.cell(row=5, column=1, value="Offres au Maroc").font = Font(size=11)
    ws.cell(row=5, column=2, value=sum(1 for r in rows if r["country"] == "Maroc"))
    ws.cell(row=6, column=1, value="Offres hors Maroc (remote)").font = Font(size=11)
    ws.cell(row=6, column=2, value=sum(1 for r in rows if r["country"] != "Maroc"))
    ws.cell(row=7, column=1, value="Offres junior / débutant / stage").font = Font(size=11)
    ws.cell(row=7, column=2, value=sum(1 for r in rows if r["seniority_bucket"] == "Junior / Debutant"))
    ws.cell(row=8, column=1, value="Offres exigeant de l'expérience").font = Font(size=11)
    ws.cell(row=8, column=2, value=sum(1 for r in rows if r["seniority_bucket"] == "Experimente"))
    ws.cell(row=9, column=1, value="Offres avec email de contact").font = Font(size=11)
    ws.cell(row=9, column=2, value=sum(1 for r in rows if r.get("contact_email") or r.get("company_email")))
    ws.cell(row=10, column=1, value="Offres avec nom de contact RH").font = Font(size=11)
    ws.cell(row=10, column=2, value=sum(1 for r in rows if r.get("recruiter_or_hr")))

    nxt = block("SOURCE", Counter(r["source"].split(" ; ")[0] for r in rows), 1, 12)
    nxt = block("VILLE", Counter(r["location_city"] for r in rows if r["country"] == "Maroc"), 1, nxt)
    block("ENTREPRISE (top 25)", Counter(r["company"] for r in rows if r.get("company")), 4, 12, 25)
    block("TYPE DE CONTRAT", Counter(r["contract_type"] for r in rows), 4, 41)
    return ws


def main():
    rows = load("merged")
    if not rows:
        print("no merged data"); return
    rows = sorted(rows, key=sort_key)
    maroc = [r for r in rows if r["country"] == "Maroc"]
    abroad = [r for r in rows if r["country"] != "Maroc"]
    junior = [r for r in rows if r["seniority_bucket"] in ("Junior / Debutant", "Non precise")]
    exper = [r for r in rows if r["seniority_bucket"] == "Experimente"]

    wb = Workbook(); wb.remove(wb.active)
    synth_sheet(wb, rows)
    write_sheet(wb, "TOUTES LES OFFRES", rows,
                f"TOUTES LES OFFRES ({len(rows)}) — Maroc + international remote. Colonne 'LIEN DE L'OFFRE' cliquable.")
    write_sheet(wb, "JUNIOR - SANS EXPERIENCE", junior,
                f"OFFRES ACCESSIBLES JUNIOR / DÉBUTANT / STAGE ({len(junior)}) — expérience non exigée ou non précisée.")
    write_sheet(wb, "AVEC EXPERIENCE", exper,
                f"OFFRES EXIGEANT DE L'EXPÉRIENCE ({len(exper)}) — 2 ans et plus, senior, manager, responsable.")
    write_sheet(wb, "HORS MAROC (REMOTE)", abroad,
                f"OFFRES HORS MAROC — 100% REMOTE / TÉLÉTRAVAIL ({len(abroad)}).")
    write_sheet(wb, "MAROC", maroc, f"OFFRES AU MAROC ({len(maroc)}).")

    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    wb.save(DEST)
    print(f"SAVED: {DEST}")
    print(f"  TOUTES LES OFFRES      {len(rows)}")
    print(f"  MAROC                  {len(maroc)}")
    print(f"  HORS MAROC (REMOTE)    {len(abroad)}")
    print(f"  JUNIOR                 {len(junior)}")
    print(f"  AVEC EXPERIENCE        {len(exper)}")
    print(f"  size: {os.path.getsize(DEST)/1024:.1f} KB")

if __name__ == "__main__":
    main()
