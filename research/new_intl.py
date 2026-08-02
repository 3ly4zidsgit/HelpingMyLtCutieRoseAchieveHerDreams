import sys, io, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from common import load, norm
import remote_curation as RC

base = load("merged"); ww = load("worldwide")
intl = [r for r in base if r["country"] != "Maroc"] + ww

reviewed_ids = set(RC.KEEP)
reviewed_board = set(RC.KEEP_BOARD)
# every offer previously present in the old workbook counted as reviewed
prev = set()
try:
    from openpyxl import load_workbook
    wb = load_workbook(r"C:\Users\pc\Desktop\Emploi\backup\Offres_Emploi_Genie_Industriel_Lean_2026_2026-07-30_1600.xlsx")
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            for c in ws[r]:
                if isinstance(c.value, str) and c.value.startswith("http"):
                    prev.add(c.value.split("?")[0].rstrip("/"))
except Exception as e:
    print("prev load err", e)

new = []
for r in intl:
    u = (r.get("url") or "").split("?")[0].rstrip("/")
    m = re.search(r"-(\d{6,})$", u)
    jid = m.group(1) if m else ""
    key = (norm(r.get("job_title", "")), norm(r.get("company", "")))
    if jid in reviewed_ids or key in reviewed_board or u in prev:
        continue
    new.append(r)

print(f"international total: {len(intl)} | never reviewed: {len(new)}\n")
out = []
for i, r in enumerate(new):
    out.append({"i": i, "title": r["job_title"], "company": r["company"],
                "loc": r.get("location_city", ""), "country": r["country"],
                "url": r["url"], "src": r["source"].split(" ; ")[0],
                "evidence": (r.get("description_snippet") or "")[:700]})
json.dump(out, open("data/new_intl.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
for o in out:
    print(f"[{o['i']}] {o['title'][:62]}")
    print(f"     {o['company'][:30]} | {o['loc'][:26]} | {o['country'][:24]} | {o['src'][:18]}")
    print(f"     {o['evidence'][:330]}")
    print()
