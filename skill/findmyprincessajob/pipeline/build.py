"""Merge -> classify -> (LLM remote verdicts) -> Excel, appending to any
workbook that already exists so the list only ever grows."""
import os, re, json, shutil, time
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from core import norm, dedupe, unmojibake, strip_accents

# ------------------------------------------------------------------ dates
FR_MONTH = {"janvier": 1, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
            "juillet": 7, "aout": 8, "septembre": 9, "octobre": 10, "novembre": 11,
            "decembre": 12}


def parse_date(s, today=None):
    today = today or datetime.now()
    t = norm(s)
    if not t:
        return None
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", t)
    if m:
        return datetime(int(m[1]), int(m[2]), int(m[3]))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", t)
    if m:
        return datetime(int(m[3]), int(m[2]), int(m[1]))
    m = re.search(r"(\d{1,2})\s+(" + "|".join(FR_MONTH) + r")\s+(\d{4})", t)
    if m:
        return datetime(int(m[3]), FR_MONTH[m[2]], int(m[1]))
    m = re.search(r"il y a (\d+)\s*(jour|semaine|mois|heure|minute)", t)
    if m:
        return today - timedelta(days={"jour": 1, "semaine": 7, "mois": 30,
                                       "heure": 0, "minute": 0}[m[2]] * int(m[1]))
    m = re.search(r"(\d+)\+?\s*(day|week|month|hour)s?\s*ago", t)
    if m:
        return today - timedelta(days={"day": 1, "week": 7, "month": 30, "hour": 0}[m[2]] * int(m[1]))
    if re.search(r"aujourd|today|just posted", t):
        return today
    if re.search(r"hier|yesterday", t):
        return today - timedelta(days=1)
    return None


# ------------------------------------------------------- seniority buckets
JUNIOR = re.compile(r"debutant|moins d.?un an|[-–]\s?1 an\b|0\s?[-a]\s?1 an|entry level|"
                    r"\bjunior\b|\bstage\b|\bstagiaire\b|\binternship\b|\bintern\b|"
                    r"fraichement diplome|sans experience|\bpfe\b|\balternance\b|"
                    r"\bapprenti|premiere experience|premier emploi|1 an d.?experience")
SENIOR = re.compile(r"\b([2-9]|1[0-9])\s*(ans|ann[ée]es|\+?\s*years?)\b|"
                    r"\bde\s*[2-9]\s*a\s*\d+\s*ans|\bsenior\b|\bconfirm[ée]\b|"
                    r"\bexp[ée]riment[ée]\b|mid[- ]senior|\bexecutive\b|\bdirector\b|"
                    r"\bdirecteur\b|\bmanager\b|\bchef de\b|\bresponsable\b|\bhead of\b|"
                    r"\blead\b|\bexpert\b|\bblack belt\b|\bprincipal\b|minimum\s*\d+\s*ans|"
                    r"\bassociate\b|\b[3-9]\s*[-a]\s*\d+\s*ans")
LI_LEVEL = {"stage": "Junior", "internship": "Junior", "debutant": "Junior",
            "entry level": "Junior", "premier emploi": "Junior", "alternance": "Junior",
            "temporaire": "Junior", "confirme": "Experimente", "associate": "Experimente",
            "mid-senior level": "Experimente", "cadre": "Experimente",
            "cadre superieur": "Experimente", "cadre dirigeant": "Experimente",
            "directeur": "Experimente", "director": "Experimente",
            "executive": "Experimente", "senior": "Experimente"}
TOP_TITLE = re.compile(r"\bvice president\b|\bvp\b|\bchief\b|\bhead of\b|\bdirector\b|"
                       r"\bdirecteur\b|\bpresident\b|\bc[ei]o\b|\bcoo\b|\bpartner\b")


# ------------------------------------------------------- experience floor
# The JUNIOR sheet exists for someone with no professional experience at all.
# What decides is therefore the FLOOR the ad sets, never the ceiling: "0 a 3 ans"
# and "de la sortie d'etudes a 6 ans" are junior, "6 mois" and "1 an minimum" are
# not. Sorting on "is there a number in the cell" puts both families on the wrong
# sheet, so every branch below reads which end of the range the number sits at.
NUM = r"(\d+(?:[.,]\d+)?)"
UNIT = r"(ans?|annees?|mois|years?|yrs?|months?)"
RANGE_RE = re.compile(NUM + r"\s*" + UNIT + r"?\s*(?:a|-|--|au?|to|jusqu.a)\s*" + NUM + r"\s*" + UNIT)
MIN_RE = re.compile(r"(?:minimum|au moins|a partir de|min\.?|at least|\+\s*de)\s*" + NUM
                    + r"\s*" + UNIT + r"|" + NUM + r"\s*" + UNIT + r"\s*(?:minimum|min\.?|et \+|ou plus)"
                    + r"|" + NUM + r"\s*\+\s*" + UNIT)
# "ouverte aux debutants" said in words - beats any ceiling figure in the cell
ZERO_RE = re.compile(r"debutant|jeune[s]? diplome|fraichement diplome|frais diplome|"
                     r"sortie d.?etudes|sortant d.?ecole|sans experience|aucune experience|"
                     r"premier emploi|premiere embauche|entry.level|new grad|graduate program|"
                     r"no experience|profil junior accepte|ouvert aux debutants|^0$")
CEIL_RE = re.compile(r"jusqu.a|maximum|\bmax\b|au plus|moins de|moins d.un|up to|\bmoins\b")
POS_QUAL_RE = re.compile(r"experience (obligatoire|exigee|requise|imperative|demandee|souhaitee|"
                         r"significative|confirmee|averee|solide|prouvee)|"
                         r"premiere experience|experience professionnelle (exigee|requise|"
                         r"obligatoire|demandee)|\bconfirme\b|\bsenior\b|\bexperimente\b|"
                         r"\bexpert\b|proven experience|obligatoire \(duree non precisee\)")
JUN_QUAL_RE = re.compile(r"\bjunior\b|\bstage\b|\bstagiaire\b|\bpfe\b|\balternance\b|\bapprenti")
MONTHS = re.compile(r"mois|months?")


def _years(value, unit):
    try:
        n = float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None
    return n / 12.0 if unit and MONTHS.search(unit) else n


def exp_floor(text):
    """Minimum experience the ad demands, in years.

    0    the ad explicitly takes someone with none
    >0   the ad demands some (0.5 marks "experience required, duration unstated")
    None the ad is silent - which is NOT the same as zero, and must never be
         read as zero: an unstated floor is unknown, not open.
    """
    t = norm(text)
    if not t:
        return None
    m = RANGE_RE.search(t)                       # a range states its own floor
    if m:
        return _years(m[1], m[2] or m[4])
    m = MIN_RE.search(t)                         # "minimum 3 ans", "3 ans et +"
    if m:
        g = [x for x in m.groups() if x]
        val = next((x for x in g if re.fullmatch(NUM, x)), None)
        unit = next((x for x in g if re.fullmatch(UNIT, x)), None)
        y = _years(val, unit)
        if y is not None:
            return y
    if re.match(r"0(\D|$)", t):                  # already normalised: "0", "0 (…)"
        return 0.0
    if ZERO_RE.search(t):                        # beats any ceiling figure
        return 0.0
    if CEIL_RE.search(t):                        # a ceiling alone leaves the floor at 0
        return 0.0
    m = re.search(NUM + r"\s*" + UNIT, t)        # a bare figure with a unit is a floor
    if m:
        return _years(m[1], m[2])
    if JUN_QUAL_RE.search(t):                    # "Junior", "profil junior"
        return 0.0
    if POS_QUAL_RE.search(t):                    # some experience, duration unstated
        return 0.5
    return None


def normalize_exp(text):
    """Task 3: an ad that takes beginners says so as a number. "Debutant accepte",
    "Jeune diplome", "Fraichement diplome" all mean the same thing and all mean 0,
    so write 0 - but only where the ad said it. A silent ad keeps its empty cell."""
    s = re.sub(r"\s+", " ", str(text or "")).strip()
    if not s or exp_floor(s) != 0:
        return s
    t = norm(s)
    if re.match(r"^0\b", t):                     # "0", "0 a 3 ans" - already explicit
        return s
    qual = ""
    m = re.search(r"\(([^)]{5,60})\)", s)        # keep a real qualifier, drop "(e)"
    if m and not re.fullmatch(r"[\W_]*\w{1,2}[\W_]*", m[1]):
        qual = f" ({m[1].strip()})"
    m = re.search(r"(?:jusqu.a|a|to|up to)\s*" + NUM + r"\s*" + UNIT, t) or \
        re.search(NUM + r"\s*" + UNIT + r"\s*(?:maximum|max)", t)
    if m:                                        # the ad gave a ceiling: keep it
        unit = "mois" if MONTHS.search(m[2] or "") else "ans"
        if m[1] in qual:                         # "Junior (2 ans maximum)" - do not say it twice
            qual = ""
        return f"0 à {m[1]} {unit}{qual}"
    return f"0{qual}"


def classify_exp(row):
    # The floor stated in the ad is the whole rule for which sheet a row lands on,
    # so it outranks every other signal - including a bucket the model locked in,
    # which used to be able to file "1 a 3 ans" as Junior.
    floor = exp_floor(row.get("experience_required", ""))
    if floor is not None:
        return "Junior" if floor <= 0 else "Experimente"
    # A bucket set by the model from the full ad text wins over the regexes
    # below: those only ever saw the listing card, which is why so many rows used
    # to come out "Non precise" while the body said "minimum 5 ans".
    if row.get("seniority_locked") and row.get("seniority_bucket"):
        return row["seniority_bucket"]
    exp = norm(row.get("experience_required", ""))
    if exp in LI_LEVEL:
        return LI_LEVEL[exp]
    if TOP_TITLE.search(norm(row.get("job_title", ""))):
        return "Experimente"
    blob = norm(" ".join([exp, row.get("job_title", ""),
                          row.get("description_snippet", "")[:400]]))
    if exp and JUNIOR.search(exp):
        return "Junior"
    if exp and SENIOR.search(exp):
        return "Experimente"
    if JUNIOR.search(blob):
        return "Junior"
    if SENIOR.search(blob):
        return "Experimente"
    return "Non precise"


def guess_contract(row):
    if row.get("contract_type"):
        return row["contract_type"]
    blob = norm(row.get("description_snippet", "") + " " + row.get("job_title", ""))
    for pat, lab in [(r"\bcdi\b", "CDI"), (r"\bcdd\b", "CDD"), (r"\bstage\b|\bpfe\b", "Stage"),
                     (r"\balternance\b|\bapprentissage\b", "Alternance"),
                     (r"\bfreelance\b", "Freelance"), (r"\binterim\b", "Interim"),
                     (r"\bfull[- ]time\b|\btemps plein\b", "Temps plein"),
                     (r"\bpart[- ]time\b|\btemps partiel\b", "Temps partiel"),
                     (r"\bcontract\b", "Contract"), (r"\bpermanent\b", "Permanent")]:
        if re.search(pat, blob):
            return lab
    return ""


MA_CITIES = ["casablanca", "tanger", "tangier", "rabat", "kenitra", "marrakech", "marrakesh",
             "agadir", "fes", "fez", "meknes", "oujda", "tetouan", "sale", "mohammedia",
             "el jadida", "safi", "nador", "berrechid", "settat", "benguerir", "laayoune",
             "khouribga", "temara", "bouskoura", "nouaceur", "skhirat", "beni mellal",
             "larache", "essaouira", "dakhla", "tiflet", "had soualem"]
DATEISH = re.compile(r"^\s*(\d+\+?\s*(jour|semaine|mois|heure|day|week|month|hour)s?|il y a\b|"
                     r".*\bago\b|aujourd|today|hier|yesterday|\d{2}/\d{2}/\d{4}|"
                     r"\d{4}-\d{2}-\d{2}|just posted)", re.I)
COUNTRY_HINT = re.compile(
    r"(etats[- ]unis|united states|usa|royaume[- ]uni|united kingdom|canada|allemagne|germany|"
    r"france|espagne|spain|portugal|italie|italy|suisse|switzerland|belgique|belgium|"
    r"pays[- ]bas|netherlands|irlande|ireland|pologne|poland|roumanie|tunisie|algerie|egypte|"
    r"senegal|nigeria|kenya|afrique du sud|emirats|emirates|qatar|turquie|inde|india|mexique|"
    r"bresil|australie|japon|chine|singapour|suede|norvege|danemark|finlande|autriche|"
    r"luxembourg|europe|worldwide|anywhere|remote|latam|emea|apac)", re.I)


def is_morocco(row):
    blob = norm(" ".join([row.get("country", ""), row.get("location_city", ""),
                          row.get("source", "")]))
    return "maroc" in blob or "morocco" in blob or any(c in blob for c in MA_CITIES)


def clean_city(row):
    raw = (row.get("location_city") or "").strip()
    t = norm(raw)
    for c in MA_CITIES:
        if c in t:
            return c.title()
    if DATEISH.match(raw):
        if not row.get("date_posted"):
            row["date_posted"] = raw
        return ""
    return re.sub(r"\s*\(.*?\)", "", raw.split(",")[0]).strip()[:60] if raw else ""


def split_location(raw):
    parts = [p.strip() for p in (raw or "").split(",") if p.strip()]
    if not parts:
        return "", ""
    country = ""
    for p in reversed(parts):
        if COUNTRY_HINT.search(strip_accents(p).lower()) or p == parts[-1]:
            country = p
            break
    return (parts[0] if parts[0] != country else ""), country


TEXT_FIELDS = ("job_title company location_city sector function description_snippet "
               "contract_type experience_required education_level recruiter_or_hr salary").split()

# ------------------------------------------------------------ closed offers
# A closure that was SEEN on the page ("cette offre n'est plus d'actualite",
# "candidatures fermees"). It is not a guess about age: "annonce publiee il y a
# 2 ans - verifier qu'elle est toujours ouverte" stays, because nobody checked it
# and an old ad can still be open.
CLOSED_RE = re.compile(r"offre expiree|offre expire|candidatures? fermees?|annonce fermee|"
                       r"n.est plus d.actualite|plus disponible|poste pourvu|offre cloturee|"
                       r"no longer accepting|position (has been )?filled|expired", re.I)


def is_closed(row):
    """True when the ad itself was read as closed. Such a row must not ship:
    the workbook is there to be applied from."""
    for f in ("deadline", "deadline_iso", "date_posted"):
        if CLOSED_RE.search(strip_accents(str(row.get(f) or ""))):
            return True
    return False


def finalize(rows):
    rows = dedupe([r for r in rows if (r.get("url") or "").startswith("http")])
    for r in rows:
        for f in TEXT_FIELDS:
            if r.get(f):
                r[f] = unmojibake(r[f])
        raw = r.get("location_city", "")
        if is_morocco(r):
            r["location_city"] = clean_city(r)
            r["country"] = "Maroc"
        else:
            city, country = split_location(raw)
            r["location_city"] = city or clean_city(r)
            r["country"] = r.get("country") or country or "International (Remote)"
        r["contract_type"] = guess_contract(r)
        r["experience_required"] = normalize_exp(r.get("experience_required", ""))
        r["seniority_bucket"] = classify_exp(r)
        d = parse_date(r.get("date_posted"))
        r["date_posted_iso"] = d.strftime("%Y-%m-%d") if d else ""
        dl = parse_date(r.get("deadline"))
        r["deadline_iso"] = dl.strftime("%Y-%m-%d") if dl else ""
    return rows


# ------------------------------------------------------------------- Excel
COLS = [("Postulé", "applied", 10),
        ("N°", "idx", 6), ("Titre du poste", "job_title", 46), ("Entreprise", "company", 26),
        ("Type de contrat", "contract_type", 17), ("Niveau d'expérience", "seniority_bucket", 17),
        ("Expérience requise", "experience_required", 20), ("Ville", "location_city", 20),
        ("Pays / Éligibilité", "country", 24), ("Secteur", "sector", 24),
        ("Fonction", "function", 24), ("Date de publication", "date_posted", 18),
        ("Date limite candidature", "deadline", 18), ("LIEN DE L'OFFRE", "url", 52),
        ("Description (extrait)", "description_snippet", 80)]
REMOTE_COL = ("Preuve 100% remote / sans visa", "remote_reason", 60)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT = PatternFill("solid", fgColor="EEF3FA")
LINK = Font(color="0563C1", underline="single", size=10)
CELL = Font(size=10)
_t = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_t, right=_t, top=_t, bottom=_t)
BUCKET = {"Junior": PatternFill("solid", fgColor="C6EFCE"),
          "Experimente": PatternFill("solid", fgColor="FCE4D6"),
          "Non precise": PatternFill("solid", fgColor="F2F2F2")}

# --------------------------------------------------------------- "Postule"
# A tick she puts there herself. It has to survive every later build, so it is
# read back out of the workbook like any other column - and it is the ONLY column
# the pipeline never writes a value into. Ticking greys the whole line out live in
# Excel (conditional formatting), so an offer she has already applied to stops
# competing for her attention without disappearing from the list.
TICK = "✔"
APPLIED_FILL = PatternFill("solid", fgColor="1F3864")
APPLIED_FONT = Font(bold=True, color="FFFFFF", size=12)
DONE_FILL = PatternFill("solid", fgColor="D9D9D9")
DONE_FONT = Font(color="808080", size=10, italic=True)


def is_applied(v):
    return norm(v).strip() in ("v", "x", "oui", "yes", "ok", "fait", "1", "true") or TICK in str(v or "")


def _clean(v):
    return re.sub(r"\s+", " ", re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", str(v or ""))).strip()


def sort_key(r):
    ts = 0.0
    d = r.get("date_posted_iso") or ""
    if d:
        try:
            ts = datetime.strptime(d, "%Y-%m-%d").timestamp()
        except ValueError:
            ts = 0.0
    return (0 if ts else 1, -ts, -r.get("score", 0))


def write_sheet(wb, name, rows, note, cols):
    ws = wb.create_sheet(name[:31])
    ws.cell(row=1, column=1, value=note).font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.row_dimensions[1].height = 22
    for c, (lab, _, w) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=lab)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 32
    for i, row in enumerate(sorted(rows, key=sort_key), 1):
        rr = i + 2
        for c, (lab, key, _) in enumerate(cols, 1):
            val = i if key == "idx" else _clean(row.get(key, ""))
            cell = ws.cell(row=rr, column=c, value=val)
            cell.font, cell.border = CELL, BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=key in ("job_title", "description_snippet",
                                                         "remote_reason", "sector"),
                                       horizontal="center" if key == "idx" else "left")
            if i % 2 == 0 and key not in ("seniority_bucket", "applied"):
                cell.fill = ALT
            if key == "seniority_bucket":
                cell.fill = BUCKET.get(row.get("seniority_bucket"), ALT)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if key == "applied":
                cell.fill, cell.font = APPLIED_FILL, APPLIED_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "url" and str(val).startswith("http"):
                cell.hyperlink, cell.font = str(val), LINK
        ws.row_dimensions[rr].height = 30
    last, wide = len(rows) + 2, get_column_letter(len(cols))
    if rows:
        ws.auto_filter.ref = f"A2:{wide}{last}"
        # tick from a dropdown so the mark is always the same character, and grey
        # the whole line out the moment it is ticked
        dv = DataValidation(type="list", formula1=f'"{TICK}"', allow_blank=True,
                            showDropDown=False)
        dv.prompt, dv.promptTitle = "Cocher une fois la candidature envoyee", "Postule"
        ws.add_data_validation(dv)
        dv.add(f"A3:A{last}")
        ws.conditional_formatting.add(
            f"A3:{wide}{last}",
            FormulaRule(formula=[f'$A3="{TICK}"'], fill=DONE_FILL, font=DONE_FONT,
                        stopIfTrue=False))
    ws.freeze_panes = ws.cell(row=3, column=2)   # header + the "Postule" column
    return ws


SHEETS = [("MAROC - JUNIOR", "Maroc", "Junior"),
          ("MAROC - AVEC EXPERIENCE", "Maroc", "Experimente"),
          ("REMOTE - JUNIOR", "Remote", "Junior"),
          ("REMOTE - AVEC EXPERIENCE", "Remote", "Experimente")]


def read_existing(path):
    """Pull previously-saved rows back out of the workbook so a re-run adds to
    them instead of replacing them."""
    if not path or not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path)
    except Exception:
        return []
    out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        hdr = [c.value for c in ws[2]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        label2key = {lab: key for lab, key, _ in COLS + [REMOTE_COL]}
        for r in range(3, ws.max_row + 1):
            vals = [c.value for c in ws[r]]
            if not any(vals):
                continue
            row = {"score": 0}
            for lab, i in idx.items():
                key = label2key.get(lab)
                if key and key != "idx" and i < len(vals):
                    row[key] = _clean(vals[i])
            if row.get("url"):
                row["source"] = row.get("source") or "(déjà présent)"
                row.setdefault("seniority_bucket",
                               "Junior" if "JUNIOR" in sname else "Experimente")
                # rows already on a REMOTE sheet were vetted in an earlier run;
                # keep that verdict or they would silently drop out on append
                if sname.upper().startswith("REMOTE"):
                    row["remote_verdict"] = "OK"
                out.append(row)
    print(f"  existing workbook: {len(out)} rows recovered", flush=True)
    return out


def publish(dest, publish_dir):
    """Drop a copy into a synced cloud folder (Google Drive Desktop, OneDrive...).

    An e-mail attachment can never be updated once sent. A file inside a synced
    folder can: overwriting it here re-syncs it, and whoever holds the share link
    always sees the current version. So the workbook is shared once, as a link,
    and every later build refreshes it in place."""
    if not publish_dir:
        return
    try:
        # Never makedirs blindly. On Google Drive's virtual filesystem the walk up
        # to the drive root raises WinError 3 even when the folder is there - and
        # the mount itself comes and goes, so isdir can be False for a moment on a
        # folder that exists. Creating anything in that state would be wrong, so a
        # missing parent means "skip and say so", never "build the tree".
        if not os.path.isdir(publish_dir):
            # the Drive mount flaps: the same path answers False, then True a few
            # seconds later. Ask again before concluding it is gone.
            for _ in range(5):
                time.sleep(2)
                if os.path.isdir(publish_dir):
                    break
        if not os.path.isdir(publish_dir):
            parent = os.path.dirname(publish_dir.rstrip("/\\"))
            if parent and not os.path.isdir(parent):
                print(f"  publication ignoree: {publish_dir} inaccessible "
                      f"(disque non monte ?) - le classeur du bureau est a jour", flush=True)
                return
            os.makedirs(publish_dir, exist_ok=True)
        target = os.path.join(publish_dir, os.path.basename(dest))
        shutil.copy2(dest, target)
        print(f"  publie -> {target}", flush=True)
    except Exception as e:
        print(f"  publication ignoree ({type(e).__name__}: {e})", flush=True)


def build(rows, dest, previous=None, publish_dir=None):
    rows = finalize(list(previous or []) + list(rows))
    remote = [r for r in rows if r["country"] != "Maroc" and r.get("remote_verdict") == "OK"]
    maroc = [r for r in rows if r["country"] == "Maroc"]
    groups = {("Maroc", "Junior"): [r for r in maroc if r["seniority_bucket"] != "Experimente"],
              ("Maroc", "Experimente"): [r for r in maroc if r["seniority_bucket"] == "Experimente"],
              ("Remote", "Junior"): [r for r in remote if r["seniority_bucket"] != "Experimente"],
              ("Remote", "Experimente"): [r for r in remote if r["seniority_bucket"] == "Experimente"]}
    wb = Workbook(); wb.remove(wb.active)
    for name, zone, bucket in SHEETS:
        data = groups[(zone, bucket)]
        cols = COLS + ([REMOTE_COL] if zone == "Remote" else [])
        label = "JUNIOR / SANS EXPÉRIENCE" if bucket == "Junior" else "AVEC EXPÉRIENCE"
        where = "MAROC" if zone == "Maroc" else "100% REMOTE MONDIAL (sans visa)"
        write_sheet(wb, name, data, f"{where} — {label} ({len(data)} offres)", cols)
    os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
    wb.save(dest)
    total = sum(len(v) for v in groups.values())
    print(f"\nSAVED: {dest}")
    for (zone, bucket), v in groups.items():
        print(f"  {zone:7s} {bucket:12s} {len(v):5d}")
    print(f"  TOTAL {total}  ({os.path.getsize(dest)/1024:.1f} KB)")
    publish(dest, publish_dir)
    return total
